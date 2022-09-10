import sys
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment

text = ""

if len(sys.argv) >= 2:
    fileName = " ".join(sys.argv[1:])
    with open(fileName, "r", encoding = "utf-8") as fil:
        text = fil.read()
else:
    print("Please execute with a text file.")
    sys.quit()

def formatCutter(txt, lis, splitter = "\n"):
    result = [] #each segments
    for i in range(1, len(lis)):
        t_ = txt.split(lis[i])[0] #entire text before the latter match
        result.append(t_)
        txt = lis[i] + lis[i].join(txt.split(lis[i])[1:]) #text excluding t_
        if txt[:len(splitter)] == splitter:
            txt = txt[len(splitter):] #clear splitter in the beginning before next step
    result.append(txt)
    return result

R = re.findall("\n\n.* \d\d:\d\d \wM", "\n\n" + text)
paragraphs = formatCutter(text, R, "\n\n") #each threads
data = []

for para in paragraphs:
    Rp = re.findall("\n.*\d\d:\d\d \wM", "\n" + para)
    sentences = formatCutter(para, Rp) #each comments in a thread
    d_ = []
    for i in range(len(sentences)):
        Rd = re.findall("\d\d:\d\d \wM", sentences[i])[0]
        bit = {}
        bit["type"] = "Q" if i == 0 else "A"
        bit["user"] = sentences[i].split(" " + Rd)[0]
        bit["time"] = Rd
        bit["contents"] = "\n".join(sentences[i].split("\n")[1:])
        d_.append(bit)
    data.append(d_)

wb = Workbook()
ws1 = wb.active
ws1.column_dimensions["B"].width = 25
ws1.column_dimensions["C"].width = 10
ws1.column_dimensions["D"].width = 300
row = ws1.max_row
for dat in data:
    for dt in dat:
        ws1.cell(row, 1, dt["type"])
        ws1.cell(row, 1).alignment = Alignment(horizontal = "center")
        ws1.cell(row, 2, dt["user"])
        ws1.cell(row, 3, dt["time"])
        ws1.cell(row, 4, dt["contents"])
        row = ws1.max_row + 1
    row = ws1.max_row + 2

wb.save(" ".join(sys.argv[1:]).replace(".txt", ".xlsx"))